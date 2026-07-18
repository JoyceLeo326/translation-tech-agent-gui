from __future__ import annotations

import csv
import json
import os
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterable


GROUPS: dict[str, tuple[str, str, str]] = {
    "A": ("A_image_translation", "A 组：图文翻译与回填", "图片 OCR、扩散修补、图文回填、Excel 清单"),
    "B": ("B_terms_style", "B 组：术语库与儿童文学风格控制", "文化术语库、风格提示词、Coze 工作流"),
    "C": ("C_text_audio_translation", "C 组：文本、DOCX 与音频翻译", "普通文本、DOCX、音视频翻译与语音合成"),
}

EXCLUDED_DIR_NAMES = {
    ".git",
    ".pytest_cache",
    ".venv",
    "__pycache__",
    "build",
    "dist",
    "runtime_outputs",
    "scratch",
    "temp",
}

ARCHIVE_SUFFIXES = {".zip", ".rar", ".7z"}
DOCUMENT_SUFFIXES = {".doc", ".docx", ".md", ".pdf", ".txt"}
SPREADSHEET_SUFFIXES = {".csv", ".xls", ".xlsx"}
IMAGE_SUFFIXES = {".bmp", ".gif", ".jpeg", ".jpg", ".png", ".svg", ".webp"}
AUDIO_SUFFIXES = {".aac", ".flac", ".m4a", ".mp3", ".wav"}
VIDEO_SUFFIXES = {".mp4", ".mov", ".avi", ".mkv"}
WORKFLOW_SUFFIXES = {".json", ".yaml", ".yml"}
CODE_SUFFIXES = {".py", ".ps1", ".sh", ".toml", ".lock"}


@dataclass(frozen=True)
class FileAsset:
    group_key: str
    group_name: str
    relative_path: str
    category: str
    suffix: str
    size_bytes: int
    modified_at: str


@dataclass(frozen=True)
class GroupSummary:
    key: str
    name: str
    description: str
    relative_path: str
    exists: bool
    file_count: int
    total_size_bytes: int
    latest_modified_at: str
    status: str
    recommendation: str
    categories: dict[str, int] = field(default_factory=dict)


@dataclass(frozen=True)
class TerminologyStats:
    count: int
    source_path: str
    sample_terms: tuple[str, ...]


@dataclass(frozen=True)
class CollaborationScan:
    project_root: Path
    scanned_at: str
    groups: tuple[GroupSummary, ...]
    assets: tuple[FileAsset, ...]
    terminology: TerminologyStats


@dataclass(frozen=True)
class OutputBundle:
    summary_markdown: str
    markdown_path: Path
    csv_path: Path
    excel_path: Path | None


def find_project_root(start: Path | None = None) -> Path:
    """Find the repository or packaged data root that contains collaboration/."""
    for candidate in _candidate_roots(start):
        if (candidate / "collaboration" / "groups").is_dir():
            return candidate
    return (start or Path.cwd()).resolve()


def output_dir_for(project_root: Path) -> Path:
    repo_output = project_root / "collaboration" / "integration" / "final_outputs" / "generated"
    try:
        repo_output.mkdir(parents=True, exist_ok=True)
        probe = repo_output / ".write-test"
        probe.write_text("ok", encoding="utf-8")
        probe.unlink(missing_ok=True)
        return repo_output
    except OSError:
        fallback = Path.home() / "Documents" / "TranslationTechAgentGUI" / "final_outputs"
        fallback.mkdir(parents=True, exist_ok=True)
        return fallback


def scan_collaboration(project_root: Path | None = None) -> CollaborationScan:
    root = find_project_root(project_root)
    scanned_at = _now_stamp()
    assets: list[FileAsset] = []
    summaries: list[GroupSummary] = []

    for key, (directory, display_name, description) in GROUPS.items():
        group_root = root / "collaboration" / "groups" / directory
        group_files = list(_iter_files(group_root)) if group_root.exists() else []
        categories: dict[str, int] = {}
        total_size = 0
        latest_mtime = 0.0

        for file_path in group_files:
            stat = file_path.stat()
            total_size += stat.st_size
            latest_mtime = max(latest_mtime, stat.st_mtime)
            category = categorize_file(file_path)
            categories[category] = categories.get(category, 0) + 1
            assets.append(
                FileAsset(
                    group_key=key,
                    group_name=display_name,
                    relative_path=_relative_to_root(file_path, root),
                    category=category,
                    suffix=file_path.suffix.lower() or "(none)",
                    size_bytes=stat.st_size,
                    modified_at=_format_mtime(stat.st_mtime),
                )
            )

        status, recommendation = _group_status(root, key, group_files)
        summaries.append(
            GroupSummary(
                key=key,
                name=display_name,
                description=description,
                relative_path=_relative_to_root(group_root, root),
                exists=group_root.exists(),
                file_count=len(group_files),
                total_size_bytes=total_size,
                latest_modified_at=_format_mtime(latest_mtime) if latest_mtime else "无",
                status=status,
                recommendation=recommendation,
                categories=dict(sorted(categories.items())),
            )
        )

    return CollaborationScan(
        project_root=root,
        scanned_at=scanned_at,
        groups=tuple(summaries),
        assets=tuple(sorted(assets, key=lambda asset: (asset.group_key, asset.category, asset.relative_path))),
        terminology=_terminology_stats(root),
    )


def load_terms(project_root: Path | None = None) -> list[dict[str, object]]:
    root = find_project_root(project_root)
    json_path = root / "collaboration" / "shared" / "terminology" / "zhonghua_culture_terms.normalized.json"
    csv_path = root / "collaboration" / "shared" / "terminology" / "zhonghua_culture_terms.normalized.csv"
    records: list[dict[str, object]] = []

    if json_path.exists():
        data = json.loads(json_path.read_text(encoding="utf-8"))
        if isinstance(data, list):
            records = [dict(item) for item in data if isinstance(item, dict)]
    elif csv_path.exists():
        with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
            records = [dict(row) for row in csv.DictReader(handle)]

    supplement = root / "collaboration" / "shared" / "corpus" / "official_culture_terms_20260718.json"
    if supplement.exists():
        data = json.loads(supplement.read_text(encoding="utf-8"))
        if isinstance(data, list):
            seen = {str(item.get("术语", "")).strip() for item in records}
            for raw in data:
                if not isinstance(raw, dict):
                    continue
                term = str(raw.get("术语", "")).strip()
                if not term or term in seen:
                    continue
                item = dict(raw)
                item.setdefault("出处页码", item.get("类别", "官方补充"))
                child_phrase = str(item.get("儿童表达", "")).strip()
                context = str(item.get("上下文片段", "")).strip()
                if child_phrase:
                    context = f"{context} 儿童表达：{child_phrase}".strip()
                item["上下文片段"] = context
                item["source_row"] = f"official-{len(records) + 1}"
                records.append(item)
                seen.add(term)

    return records


def search_terms(query: str, project_root: Path | None = None, limit: int = 30) -> list[dict[str, object]]:
    terms = load_terms(project_root)
    query = query.strip().lower()
    if not query:
        return terms[:limit]

    scored: list[tuple[int, dict[str, object]]] = []
    for record in terms:
        term = str(record.get("术语", ""))
        english = str(record.get("英文翻译", ""))
        context = str(record.get("上下文片段", ""))
        category = str(record.get("类别", ""))
        source = str(record.get("来源", ""))
        haystack = f"{term} {english} {context} {category} {source}".lower()
        score = 0
        if term and term in query:
            score += 5
        if query in term.lower() or query in english.lower():
            score += 4
        if query in haystack:
            score += 2
        if score:
            scored.append((score, record))

    scored.sort(key=lambda item: (-item[0], str(item[1].get("source_row", ""))))
    return [record for _, record in scored[:limit]]


def format_terms_markdown(records: Iterable[dict[str, object]], title: str = "术语检索结果") -> str:
    rows = list(records)
    if not rows:
        return f"# {title}\n\n未检索到匹配术语。"

    lines = [
        f"# {title}",
        "",
        "| 术语 | 英文翻译 | 出处页码 | 上下文片段 |",
        "| --- | --- | --- | --- |",
    ]
    for record in rows:
        lines.append(
            "| {term} | {english} | {page} | {context} |".format(
                term=_cell(record.get("术语", "")),
                english=_cell(record.get("英文翻译", "")),
                page=_cell(record.get("出处页码", "")),
                context=_cell(str(record.get("上下文片段", ""))[:120]),
            )
        )
    return "\n".join(lines)


def format_dashboard_markdown(scan: CollaborationScan) -> str:
    category_counts: dict[str, int] = {}
    total_size = 0
    for asset in scan.assets:
        category_counts[asset.category] = category_counts.get(asset.category, 0) + 1
        total_size += asset.size_bytes
    ready_root = scan.project_root / "collaboration" / "integration" / "final_outputs" / "ready_to_use"
    ready_files = list(_iter_files(ready_root)) if ready_root.exists() else []
    lines = [
        "# 译述 YISHU · 多模态外译统一资源报告",
        "",
        f"- 扫描时间：{scan.scanned_at}",
        f"- 已连接资源：{len(scan.assets)} 个，总大小 {format_size(total_size)}",
        f"- 文化术语与官方补充：{scan.terminology.count} 条",
        f"- 统一成品区：{len(ready_files)} 个可直接使用文件",
        "",
        "## 内容类型",
        "",
    ]
    lines.extend(f"- {category}：{count}" for category, count in sorted(category_counts.items()))

    if scan.terminology.sample_terms:
        lines.extend(["", "## 术语样例", ""])
        lines.append("、".join(scan.terminology.sample_terms))

    lines.extend(
        [
            "",
            "## 统一生产链",
            "",
            "1. 图文、DOCX 或音视频素材先进入统一素材生产页。",
            "2. 智能翻译自动加载文化术语与儿童文学风格约束。",
            "3. 人工审核列拥有最终优先级，确认后再执行文档回填或英文配音。",
            "4. 新成品进入 `generated/`；稳定初始成品位于 `ready_to_use/`。",
        ]
    )
    return "\n".join(lines)


def run_group_adapter(group_key: str, prompt: str = "", project_root: Path | None = None) -> str:
    root = find_project_root(project_root)
    group_key = group_key.upper()
    scan = scan_collaboration(root)
    prompt = prompt.strip()

    if group_key == "A":
        return _run_a_adapter(root, scan)
    if group_key == "B":
        terms = search_terms(prompt, root, limit=20) if prompt else load_terms(root)[:20]
        return "\n\n".join(
            [
                "# 文化术语与儿童文学风格约束",
                "统一术语库已直接参与智能翻译；输入文本包含文化术语时优先采用已审核译法。",
                format_terms_markdown(terms, "术语命中"),
                "儿童文学风格规则已接入统一智能翻译流程。",
            ]
        )
    if group_key == "C":
        return _run_c_adapter(root, scan)
    return "该兼容入口已迁移至统一素材生产页。"


def write_integration_outputs(scan: CollaborationScan, prompt: str = "") -> OutputBundle:
    out_dir = output_dir_for(scan.project_root)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    markdown_path = out_dir / f"integration_summary_{timestamp}.md"
    csv_path = out_dir / f"integration_assets_{timestamp}.csv"
    excel_path = out_dir / f"integration_report_{timestamp}.xlsx"

    summary = format_dashboard_markdown(scan)
    if prompt.strip():
        summary = f"{summary}\n\n## 本次整合目标\n\n{prompt.strip()}\n"
    markdown_path.write_text(summary, encoding="utf-8")

    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["category", "suffix", "size_bytes", "modified_at", "relative_path"])
        for asset in scan.assets:
            writer.writerow(
                [
                    asset.category,
                    asset.suffix,
                    asset.size_bytes,
                    asset.modified_at,
                    asset.relative_path,
                ]
            )

    try:
        _write_excel_report(excel_path, scan)
        actual_excel_path: Path | None = excel_path
    except ImportError:
        actual_excel_path = None

    return OutputBundle(
        summary_markdown=_format_output_bundle(summary, markdown_path, csv_path, actual_excel_path),
        markdown_path=markdown_path,
        csv_path=csv_path,
        excel_path=actual_excel_path,
    )


def categorize_file(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in ARCHIVE_SUFFIXES:
        return "压缩包"
    if suffix in DOCUMENT_SUFFIXES:
        return "文档"
    if suffix in SPREADSHEET_SUFFIXES:
        return "表格"
    if suffix in IMAGE_SUFFIXES:
        return "图片"
    if suffix in AUDIO_SUFFIXES:
        return "音频"
    if suffix in VIDEO_SUFFIXES:
        return "视频"
    if suffix in WORKFLOW_SUFFIXES:
        return "工作流/结构化数据"
    if suffix in CODE_SUFFIXES:
        return "代码/配置"
    return "其它"


def format_size(size_bytes: int) -> str:
    value = float(size_bytes)
    for unit in ("B", "KB", "MB", "GB"):
        if value < 1024 or unit == "GB":
            return f"{value:.1f} {unit}" if unit != "B" else f"{int(value)} B"
        value /= 1024
    return f"{value:.1f} GB"


def _candidate_roots(start: Path | None) -> list[Path]:
    raw: list[Path] = []
    if start is not None:
        raw.append(start)
    raw.append(Path.cwd())
    raw.append(Path(__file__).resolve().parents[2])

    env_root = os.getenv("TRANSLATION_GUI_PROJECT_ROOT")
    if env_root:
        raw.append(Path(env_root))

    frozen_root = getattr(sys, "_MEIPASS", None)
    if frozen_root:
        raw.append(Path(frozen_root))

    if getattr(sys, "frozen", False):
        exe_parent = Path(sys.executable).resolve().parent
        raw.extend([exe_parent, exe_parent.parent])

    candidates: list[Path] = []
    seen: set[str] = set()
    for item in raw:
        try:
            resolved = item.resolve()
        except OSError:
            continue
        for candidate in (resolved, *resolved.parents):
            key = str(candidate).lower()
            if key not in seen:
                seen.add(key)
                candidates.append(candidate)
    return candidates


def _iter_files(root: Path) -> Iterable[Path]:
    if not root.exists():
        return
    for file_path in root.rglob("*"):
        if not file_path.is_file():
            continue
        if any(part in EXCLUDED_DIR_NAMES for part in file_path.parts):
            continue
        yield file_path


def _group_status(root: Path, key: str, files: list[Path]) -> tuple[str, str]:
    if not files:
        return "未接入", "来源目录为空"

    if key == "A":
        note = (
            root
            / "collaboration"
            / "groups"
            / "A_image_translation"
            / "deliverables"
            / "notes"
            / "A组第二次提交复核记录_20260717.md"
        )
        if note.exists():
            return "已接入", "图文终版、审校清单与 17 页预览已进入统一生产链"
        return "已归档", "来源材料已保留用于追溯"

    if key == "B":
        terms = root / "collaboration" / "shared" / "terminology" / "zhonghua_culture_terms.normalized.json"
        if terms.exists():
            return "已接入", "作为统一术语约束和风格控制输入"
        return "未接入", "共享术语文件不存在"

    if key == "C":
        revised_docx = root / "collaboration" / "groups" / "C_text_audio_translation" / "deliverables" / "docx_translation" / "revised_20260717"
        revised_audio = root / "collaboration" / "groups" / "C_text_audio_translation" / "deliverables" / "audio_video_workflow" / "revised_20260717"
        if revised_docx.exists() and revised_audio.exists():
            return "已接入", "DOCX 与音视频能力已进入统一生产链"
        return "已归档", "来源材料已保留用于追溯"

    return "已归档", "可按清单接入"


def _terminology_stats(root: Path) -> TerminologyStats:
    terms = load_terms(root)
    source = root / "collaboration" / "shared" / "terminology" / "zhonghua_culture_terms.normalized.json"
    supplement = root / "collaboration" / "shared" / "corpus" / "official_culture_terms_20260718.json"
    samples = tuple(str(record.get("术语", "")) for record in terms[:12] if record.get("术语"))
    source_label = _relative_to_root(source, root) if source.exists() else "未找到"
    if supplement.exists():
        source_label = f"{source_label} + {_relative_to_root(supplement, root)}"
    return TerminologyStats(
        count=len(terms),
        source_path=source_label,
        sample_terms=samples,
    )


def _run_a_adapter(root: Path, scan: CollaborationScan) -> str:
    group_root = root / "collaboration" / "groups" / "A_image_translation"
    latest = group_root / "deliverables" / "extracted_20260717_update"
    manifest = latest / "manifests" / "translation_manifest_reviewed.xlsx"
    final_docx = latest / "final_outputs" / "翻译资源编写-中国文化知识百科_A组更新完整修正版.docx"
    preview = latest / "previews" / "final_docx_pages_contact_sheet.jpg"
    validation = latest / "validation" / "validation_report.json"
    summary = next(group for group in scan.groups if group.key == "A")
    return "\n".join(
        [
            "# 图文翻译与版面回填",
            "",
            f"- 已接入来源文件：{summary.file_count}",
            f"- 71 条审校清单：`{_relative_to_root(manifest, root)}`",
            f"- 最终 DOCX：`{_relative_to_root(final_docx, root)}`",
            f"- 17 页渲染总览：`{_relative_to_root(preview, root)}`",
            f"- 机器可读校验：`{_relative_to_root(validation, root)}`",
            "",
            "以上成品已纳入统一素材生产、资源检索和交付中心。",
        ]
    )


def _run_c_adapter(root: Path, scan: CollaborationScan) -> str:
    group_root = root / "collaboration" / "groups" / "C_text_audio_translation"
    docx_revised = group_root / "deliverables" / "docx_translation" / "revised_20260717"
    audio_revised = group_root / "deliverables" / "audio_video_workflow" / "revised_20260717"
    summary = next(group for group in scan.groups if group.key == "C")
    return "\n".join(
        [
            "# 文本、DOCX 与音视频生产通道",
            "",
            f"- 已接入来源文件：{summary.file_count}",
            f"- DOCX 生产资源：`{_relative_to_root(docx_revised, root)}`",
            f"- 音视频生产资源：`{_relative_to_root(audio_revised, root)}`",
            "",
            "DOCX 提取、审校、回填以及音频审校、语音合成均已由统一生产页直接调用。",
        ]
    )


def _write_excel_report(path: Path, scan: CollaborationScan) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "统一概览"
    summary_sheet.append(["内容类型", "文件数", "总大小Bytes"])
    categories: dict[str, tuple[int, int]] = {}
    for asset in scan.assets:
        count, size = categories.get(asset.category, (0, 0))
        categories[asset.category] = (count + 1, size + asset.size_bytes)
    for category, (count, size) in sorted(categories.items()):
        summary_sheet.append([category, count, size])

    assets_sheet = workbook.create_sheet("资产清单")
    assets_sheet.append(["类别", "后缀", "大小Bytes", "修改时间", "路径"])
    for asset in scan.assets:
        assets_sheet.append(
            [
                asset.category,
                asset.suffix,
                asset.size_bytes,
                asset.modified_at,
                asset.relative_path,
            ]
        )

    term_sheet = workbook.create_sheet("术语样例")
    term_sheet.append(["术语", "英文翻译", "出处页码", "上下文片段"])
    for record in load_terms(scan.project_root)[:80]:
        term_sheet.append(
            [
                record.get("术语", ""),
                record.get("英文翻译", ""),
                record.get("出处页码", ""),
                record.get("上下文片段", ""),
            ]
        )

    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2F6FED")
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 60)

    workbook.save(path)


def _format_output_bundle(summary: str, markdown_path: Path, csv_path: Path, excel_path: Path | None) -> str:
    lines = [
        summary,
        "",
        "## 已生成文件",
        "",
        f"- Markdown 总结：`{markdown_path}`",
        f"- CSV 资产清单：`{csv_path}`",
    ]
    if excel_path is not None:
        lines.append(f"- Excel 整合报告：`{excel_path}`")
    else:
        lines.append("- Excel 整合报告：未生成，当前环境缺少 openpyxl")
    return "\n".join(lines)


def _relative_to_root(path: Path, root: Path) -> str:
    try:
        return path.relative_to(root).as_posix()
    except ValueError:
        return str(path)


def _format_mtime(timestamp: float) -> str:
    return datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d %H:%M:%S")


def _now_stamp() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _cell(value: object) -> str:
    return str(value).replace("|", "\\|").replace("\r", " ").replace("\n", " ").strip()
