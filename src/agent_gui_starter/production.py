from __future__ import annotations

import base64
import json
import re
import subprocess
import tempfile
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Callable, Iterable, Sequence

from docx import Document
from docx.document import Document as DocumentType
from docx.table import Table, _Cell
from docx.text.paragraph import Paragraph
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .integration import find_project_root, output_dir_for


W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
W = f"{{{W_NS}}}"
XML_SPACE = "{http://www.w3.org/XML/1998/namespace}space"
CHINESE_RE = re.compile(r"[\u3400-\u9fff]")
SENTENCE_SPLIT_RE = re.compile(r"(?<=[。！？!?])\s*|\n+")

SOURCE_ALIASES = ("中文原文", "书本文字", "音频文字", "原文", "source", "chinese", "cn")
MACHINE_ALIASES = ("机器英文译文", "机器英文", "机器译文", "机器翻译", "英文译文", "machine", "mt")
REVIEW_ALIASES = ("人工审核", "人工审校", "人工校对", "审核译文", "审校译文", "final", "review", "manual")


@dataclass(frozen=True)
class ReviewRow:
    index: int
    source_text: str
    machine_translation: str
    reviewed_translation: str
    location: str = ""

    @property
    def final_translation(self) -> str:
        return self.reviewed_translation.strip() or self.machine_translation.strip()


@dataclass(frozen=True)
class ProductionResult:
    title: str
    summary: str
    artifacts: tuple[Path, ...]
    metrics: dict[str, int | str]

    def to_markdown(self) -> str:
        lines = [f"# {self.title}", "", self.summary.strip(), "", "## 处理结果", ""]
        for label, value in self.metrics.items():
            lines.append(f"- {label}：{value}")
        if self.artifacts:
            lines.extend(["", "## 可直接使用的文件", ""])
            lines.extend(f"- `{path}`" for path in self.artifacts)
        return "\n".join(lines)

    def to_payload(self, kind: str) -> str:
        return json.dumps(
            {
                "kind": kind,
                "message": self.to_markdown(),
                "artifacts": [str(path) for path in self.artifacts],
                "metrics": self.metrics,
            },
            ensure_ascii=False,
        )


def extract_docx_to_review(
    source_docx: Path | str,
    output_path: Path | str | None = None,
) -> ProductionResult:
    source = _require_file(source_docx, ".docx")
    segments = extract_docx_segments(source)
    if not segments:
        raise ValueError("DOCX 中没有可提取的正文、表格或页眉页脚文字。")

    output = Path(output_path) if output_path else _timestamped_output(source, "审校表", ".xlsx")
    output.parent.mkdir(parents=True, exist_ok=True)
    _write_review_workbook(output, segments, source)
    chinese_segments = sum(bool(CHINESE_RE.search(text)) for _, text in segments)
    return ProductionResult(
        title="DOCX 文字已提取",
        summary="正文、表格、页眉和页脚已整理为统一审校表，可直接填写机器译文或人工审核列。",
        artifacts=(output,),
        metrics={"提取条目": len(segments), "含中文条目": chinese_segments, "源文件": source.name},
    )


def translate_review_workbook(
    review_workbook: Path | str,
    translator: Callable[[Sequence[str]], Sequence[str]],
    output_path: Path | str | None = None,
    batch_size: int = 24,
) -> ProductionResult:
    source = _require_file(review_workbook, ".xlsx")
    rows = load_review_rows(source)
    pending = [row for row in rows if row.source_text and not row.machine_translation]
    if not pending:
        raise ValueError("审校表中没有待翻译的条目。")

    translations: dict[int, str] = {}
    for start in range(0, len(pending), batch_size):
        batch = pending[start : start + batch_size]
        translated = [str(item).strip() for item in translator([row.source_text for row in batch])]
        if len(translated) != len(batch):
            raise ValueError(f"智能翻译返回 {len(translated)} 条，预期 {len(batch)} 条，已停止写入。")
        translations.update({row.index: text for row, text in zip(batch, translated)})

    output = Path(output_path) if output_path else _timestamped_output(source, "机器翻译", ".xlsx")
    _copy_review_with_translations(source, output, translations)
    return ProductionResult(
        title="机器译文已写入审校表",
        summary="译文已按原行号写入，人工审核列保持独立，可继续逐条审校后用于 DOCX 回填或语音合成。",
        artifacts=(output,),
        metrics={"本次翻译": len(translations), "审校表总条目": len(rows), "源文件": source.name},
    )


def refill_docx_from_review(
    source_docx: Path | str,
    review_workbook: Path | str,
    output_path: Path | str | None = None,
) -> ProductionResult:
    source = _require_file(source_docx, ".docx")
    review = _require_file(review_workbook, ".xlsx")
    rows = load_review_rows(review)
    pairs = [(row.source_text, row.final_translation) for row in rows if row.source_text and row.final_translation]
    if not pairs:
        raise ValueError("审校表没有可回填译文；请填写人工审核列，或保留机器英文译文。")

    output = Path(output_path) if output_path else _timestamped_output(source, "英文回填", ".docx")
    report_path = output.with_suffix(".回填报告.json")
    output.parent.mkdir(parents=True, exist_ok=True)

    mapping: dict[str, str] = {}
    originals: dict[str, str] = {}
    for cn_text, en_text in pairs:
        key = normalize_text(cn_text)
        if key and key not in mapping:
            mapping[key] = en_text.strip()
            originals[key] = cn_text
    sorted_keys = sorted(mapping, key=len, reverse=True)
    matched: set[str] = set()
    failed_xml: list[str] = []
    residue: list[str] = []
    replaced_blocks = 0

    with zipfile.ZipFile(source, "r") as zin, zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            name = item.filename
            process = _is_text_xml(name)
            if not process:
                zout.writestr(item, zin.read(name))
                continue
            try:
                root = ET.fromstring(zin.read(name))
                count = _replace_xml_paragraphs(
                    root,
                    sorted_keys,
                    mapping,
                    matched,
                    bool(re.match(r"word/(header|footer)\d+\.xml$", name)),
                )
                replaced_blocks += count
                for snippet in _collect_chinese_residue(root):
                    if snippet not in residue:
                        residue.append(snippet)
                zout.writestr(item, ET.tostring(root, xml_declaration=True, encoding="UTF-8"))
            except Exception:
                failed_xml.append(name)
                zout.writestr(item, zin.read(name))

    unmatched = [originals[key] for key in originals if key not in matched]
    report = {
        "source_docx": str(source),
        "review_workbook": str(review),
        "output_docx": str(output),
        "review_items": len(mapping),
        "matched_items": len(matched),
        "unmatched_items": len(unmatched),
        "replaced_xml_blocks": replaced_blocks,
        "failed_xml_files": failed_xml,
        "chinese_residue_samples": residue[:30],
        "unmatched_samples": unmatched[:30],
        "generated_at": datetime.now().isoformat(timespec="seconds"),
    }
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    if failed_xml:
        raise RuntimeError(f"DOCX 有 {len(failed_xml)} 个文本 XML 未能处理，详情见 {report_path}")

    return ProductionResult(
        title="DOCX 英文回填完成",
        summary="译文已直接写入 Word XML，保留原文档版式、图片、字体和段落结构，并生成覆盖报告。",
        artifacts=(output, report_path),
        metrics={
            "审校条目": len(mapping),
            "命中条目": len(matched),
            "替换文本块": replaced_blocks,
            "未命中条目": len(unmatched),
        },
    )


def create_audio_review_from_transcript(
    transcript: str,
    source_name: str,
    output_path: Path | str | None = None,
) -> ProductionResult:
    segments = [("音频", text) for text in _split_sentences(transcript) if text.strip()]
    if not segments:
        raise ValueError("转写结果为空，无法生成审校表。")
    output = Path(output_path) if output_path else output_dir_for(find_project_root()) / f"{_safe_stem(source_name)}_音频审校表_{_stamp()}.xlsx"
    _write_review_workbook(output, segments, Path(source_name), audio_mode=True)
    return ProductionResult(
        title="音频转写审校表已生成",
        summary="转写文本已按句切分，后续可执行智能翻译、人工审核和英文语音合成。",
        artifacts=(output,),
        metrics={"转写条目": len(segments), "源文件": Path(source_name).name},
    )


def synthesize_audio_from_review(
    review_workbook: Path | str,
    output_path: Path | str | None = None,
) -> ProductionResult:
    review = _require_file(review_workbook, ".xlsx")
    rows = load_review_rows(review)
    final_lines = [row.final_translation for row in rows if row.final_translation]
    if not final_lines:
        raise ValueError("审校表没有可合成的英文译文。")
    output = Path(output_path) if output_path else _timestamped_output(review, "英文配音", ".wav")
    output.parent.mkdir(parents=True, exist_ok=True)
    transcript_path = output.with_suffix(".txt")
    qr_path = output.with_suffix(".二维码.png")
    transcript_path.write_text("\n".join(final_lines), encoding="utf-8")
    _synthesize_with_windows_sapi(transcript_path, output)
    _generate_local_qr(output, qr_path)
    if not output.exists() or output.stat().st_size < 1024:
        raise RuntimeError("系统语音合成没有生成有效 WAV 文件。")
    return ProductionResult(
        title="英文语音已生成",
        summary="优先采用人工审核译文，空白审核行自动沿用机器译文；同时输出朗读文本和本机播放二维码。",
        artifacts=(output, transcript_path, qr_path),
        metrics={"合成句数": len(final_lines), "音频大小": output.stat().st_size, "格式": "WAV"},
    )


def load_review_rows(path: Path | str) -> list[ReviewRow]:
    workbook_path = _require_file(path, ".xlsx")
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    sheet = workbook.active
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    source_index = _find_column(headers, SOURCE_ALIASES)
    machine_index = _find_column(headers, MACHINE_ALIASES)
    review_index = _find_column(headers, REVIEW_ALIASES)
    location_index = _find_column(headers, ("位置", "location"))
    if source_index is None:
        raise ValueError(f"审校表缺少中文原文列，当前表头：{headers}")
    if machine_index is None and review_index is None:
        raise ValueError(f"审校表缺少机器英文译文或人工审核列，当前表头：{headers}")

    rows: list[ReviewRow] = []
    for excel_row, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        source = _cell(values, source_index)
        if not source:
            continue
        rows.append(
            ReviewRow(
                index=excel_row,
                source_text=source,
                machine_translation=_cell(values, machine_index),
                reviewed_translation=_cell(values, review_index),
                location=_cell(values, location_index),
            )
        )
    workbook.close()
    return rows


def extract_docx_segments(path: Path | str) -> list[tuple[str, str]]:
    source = _require_file(path, ".docx")
    document = Document(source)
    segments: list[tuple[str, str]] = []
    paragraph_no = 0
    table_no = 0
    for block in _iter_block_items(document):
        if isinstance(block, Paragraph):
            paragraph_no += 1
            _append_text_segments(segments, f"正文 P{paragraph_no}", block.text)
        else:
            table_no += 1
            for row_no, row in enumerate(block.rows, start=1):
                for cell_no, cell in enumerate(row.cells, start=1):
                    _append_text_segments(segments, f"表格 T{table_no} R{row_no}C{cell_no}", cell.text)

    for section_no, section in enumerate(document.sections, start=1):
        for kind, container in (("页眉", section.header), ("页脚", section.footer)):
            for para_no, paragraph in enumerate(container.paragraphs, start=1):
                _append_text_segments(segments, f"{kind} S{section_no}P{para_no}", paragraph.text)
    return _deduplicate_segments(segments)


def validate_c_docx_samples(project_root: Path | str | None = None) -> ProductionResult:
    root = find_project_root(Path(project_root) if project_root else None)
    cases_root = root / "collaboration/groups/C_text_audio_translation/deliverables/docx_translation/revised_20260717/test_cases"
    if not cases_root.is_dir():
        raise FileNotFoundError(f"未找到 DOCX 测试样例：{cases_root}")
    reports: list[dict[str, object]] = []
    for case_dir in sorted(path for path in cases_root.iterdir() if path.is_dir()):
        xlsx_files = list(case_dir.glob("*.xlsx"))
        docx_files = list(case_dir.glob("*.docx"))
        if len(xlsx_files) != 1 or len(docx_files) < 2:
            reports.append({"case": case_dir.name, "passed": False, "reason": "输入、审校表或成品数量不完整"})
            continue
        review_rows = load_review_rows(xlsx_files[0])
        source = max(
            docx_files,
            key=lambda path: sum(
                1
                for _, text in extract_docx_segments(path)
                for char in text
                if "\u4e00" <= char <= "\u9fff"
            ),
        )
        output = output_dir_for(root) / f"DOCX样例_{case_dir.name}_{_stamp()}.docx"
        result = refill_docx_from_review(source, xlsx_files[0], output)
        report_data = json.loads(result.artifacts[1].read_text(encoding="utf-8"))
        passed = output.exists() and output.stat().st_size > 1000 and report_data["failed_xml_files"] == []
        reports.append(
            {
                "case": case_dir.name,
                "passed": passed,
                "review_rows": len(review_rows),
                "matched_items": report_data["matched_items"],
                "replaced_xml_blocks": report_data["replaced_xml_blocks"],
                "output": str(output),
            }
        )
    report_path = output_dir_for(root) / f"DOCX_五套样例端到端验收_{_stamp()}.json"
    report_path.write_text(json.dumps(reports, ensure_ascii=False, indent=2), encoding="utf-8")
    passed_count = sum(bool(item.get("passed")) for item in reports)
    if passed_count != len(reports):
        raise RuntimeError(f"DOCX 样例仅通过 {passed_count}/{len(reports)}，详情见 {report_path}")
    return ProductionResult(
        title="DOCX 五套样例验收通过",
        summary="每套样例都重新执行了审校表读取、Word XML 回填、成品生成和失败 XML 检查。",
        artifacts=(report_path,),
        metrics={"通过样例": f"{passed_count}/{len(reports)}", "失败 XML": 0},
    )


def normalize_text(text: str) -> str:
    value = re.sub(r"\s+", "", str(text).strip())
    table = str.maketrans("，、；。！？「」“”‘’", ",,;.!?\"\"\"\"''")
    return value.translate(table)


def _write_review_workbook(
    output: Path,
    segments: Sequence[tuple[str, str]],
    source: Path,
    audio_mode: bool = False,
) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "翻译审校"
    source_header = "音频文字" if audio_mode else "中文原文"
    headers = ["序号", "位置", source_header, "机器英文译文", "人工审核", "审校状态", "备注"]
    sheet.append(headers)
    for index, (location, text) in enumerate(segments, start=1):
        sheet.append([index, location, text, "", "", "待翻译", ""])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = {"A": 8, "B": 22, "C": 58, "D": 58, "E": 58, "F": 14, "G": 28}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="245FCC")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    info = workbook.create_sheet("使用说明")
    info_rows = [
        ("源文件", str(source)),
        ("生成时间", datetime.now().isoformat(timespec="seconds")),
        ("操作顺序", "机器翻译 -> 人工审核 -> DOCX回填或语音合成"),
        ("回填规则", "人工审核列优先；为空时自动采用机器英文译文。"),
    ]
    for key, value in info_rows:
        info.append([key, value])
    info.column_dimensions["A"].width = 18
    info.column_dimensions["B"].width = 90
    workbook.save(output)


def _copy_review_with_translations(source: Path, output: Path, translations: dict[int, str]) -> None:
    workbook = load_workbook(source)
    sheet = workbook.active
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    machine_index = _find_column(headers, MACHINE_ALIASES)
    status_index = _find_column(headers, ("审校状态", "status"))
    if machine_index is None:
        machine_index = sheet.max_column
        sheet.cell(1, machine_index + 1, "机器英文译文")
    for row_index, translation in translations.items():
        sheet.cell(row_index, machine_index + 1, translation)
        if status_index is not None:
            sheet.cell(row_index, status_index + 1, "待人工审核")
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


def _replace_xml_paragraphs(
    root: ET.Element,
    sorted_keys: Sequence[str],
    mapping: dict[str, str],
    matched: set[str],
    header_footer: bool,
) -> int:
    replaced = 0
    for paragraph in root.iter(f"{W}p"):
        text_nodes = list(paragraph.iter(f"{W}t"))
        if not text_nodes:
            continue
        full_text = "".join(node.text or "" for node in text_nodes).strip()
        if not full_text:
            continue
        normalized = normalize_text(re.sub(r"\d+", "", full_text) if header_footer else full_text)
        candidates = [normalized]
        if header_footer:
            candidates.extend([normalize_text(f"[页眉]{full_text}"), normalize_text(f"[页脚]{full_text}")])
        translation = ""
        current_matches: list[str] = []
        for candidate in candidates:
            if candidate in mapping:
                translation = mapping[candidate]
                current_matches = [candidate]
                break
            contained = [key for key in sorted_keys if key in candidate]
            if contained and sum(len(key) for key in contained) >= len(candidate) * 0.3:
                contained.sort(key=candidate.find)
                translation = " ".join(mapping[key] for key in contained)
                current_matches = contained
                break
            reverse = next((key for key in sorted_keys if candidate and candidate in key), None)
            if reverse:
                translation = mapping[reverse]
                current_matches = [reverse]
                break
        if not translation:
            continue
        for marker in ("[Header]", "[Footer]", "[页眉]", "[页脚]"):
            translation = translation.replace(marker, "").strip()
        matched.update(current_matches)
        _distribute_text(text_nodes, translation)
        replaced += 1
    return replaced


def _distribute_text(nodes: Sequence[ET.Element], translation: str) -> None:
    lengths = [len(node.text or "") for node in nodes]
    total = sum(lengths)
    if total <= 0 or len(nodes) == 1:
        nodes[0].text = translation
        nodes[0].set(XML_SPACE, "preserve")
        for node in nodes[1:]:
            node.text = ""
        return
    consumed = 0
    for node, original_length in zip(nodes, lengths):
        start = round(consumed / total * len(translation))
        consumed += original_length
        end = round(consumed / total * len(translation))
        node.text = translation[start:end]
        node.set(XML_SPACE, "preserve")


def _collect_chinese_residue(root: ET.Element, limit: int = 30) -> list[str]:
    snippets: list[str] = []
    for node in root.iter(f"{W}t"):
        value = (node.text or "").strip()
        if value and CHINESE_RE.search(value) and value[:120] not in snippets:
            snippets.append(value[:120])
        if len(snippets) >= limit:
            break
    return snippets


def _is_text_xml(name: str) -> bool:
    excluded = ("word/_rels/", "word/theme/", "word/fontTable", "word/settings", "word/styles")
    return name.startswith("word/") and name.endswith(".xml") and not name.startswith(excluded)


def _iter_block_items(parent: DocumentType | _Cell) -> Iterable[Paragraph | Table]:
    if isinstance(parent, DocumentType):
        parent_element = parent.element.body
        parent_object = parent
    elif isinstance(parent, _Cell):
        parent_element = parent._tc
        parent_object = parent
    else:  # pragma: no cover - internal guard.
        raise TypeError("unsupported parent")
    for child in parent_element.iterchildren():
        if isinstance(child, CT_P):
            yield Paragraph(child, parent_object)
        elif isinstance(child, CT_Tbl):
            yield Table(child, parent_object)


def _append_text_segments(target: list[tuple[str, str]], location: str, text: str) -> None:
    clean = re.sub(r"[\t\u00a0]+", " ", str(text or "")).strip()
    for sentence in _split_sentences(clean):
        sentence = sentence.strip()
        if sentence:
            target.append((location, sentence))


def _split_sentences(text: str) -> list[str]:
    return [part.strip() for part in SENTENCE_SPLIT_RE.split(text) if part.strip()]


def _deduplicate_segments(segments: Sequence[tuple[str, str]]) -> list[tuple[str, str]]:
    result: list[tuple[str, str]] = []
    seen: set[tuple[str, str]] = set()
    for location, text in segments:
        key = (location, text)
        if key not in seen:
            seen.add(key)
            result.append(key)
    return result


def _find_column(headers: Sequence[str], aliases: Sequence[str]) -> int | None:
    normalized = [re.sub(r"[\s_]", "", value).lower() for value in headers]
    for alias in aliases:
        alias_norm = re.sub(r"[\s_]", "", alias).lower()
        for index, value in enumerate(normalized):
            if alias_norm and alias_norm in value:
                return index
    return None


def _cell(values: Sequence[object], index: int | None) -> str:
    if index is None or index >= len(values):
        return ""
    value = values[index]
    return "" if value is None else str(value).strip()


def _require_file(path: Path | str, suffix: str) -> Path:
    value = Path(path).expanduser().resolve()
    if not value.is_file():
        raise FileNotFoundError(f"文件不存在：{value}")
    if value.suffix.lower() != suffix.lower():
        raise ValueError(f"需要 {suffix} 文件，当前选择：{value.name}")
    return value


def _timestamped_output(source: Path, label: str, suffix: str) -> Path:
    return output_dir_for(find_project_root()) / f"{_safe_stem(source.stem)}_{label}_{_stamp()}{suffix}"


def _safe_stem(value: str) -> str:
    cleaned = re.sub(r"[<>:\"/\\|?*]+", "_", value).strip(" ._")
    return cleaned[:80] or "output"


def _stamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _synthesize_with_windows_sapi(text_path: Path, output_path: Path) -> None:
    script = r"""
param([string]$TextPath, [string]$OutputPath)
$ErrorActionPreference = 'Stop'
Add-Type -AssemblyName System.Speech
$text = [System.IO.File]::ReadAllText($TextPath, [System.Text.Encoding]::UTF8)
$speaker = New-Object System.Speech.Synthesis.SpeechSynthesizer
try {
    $englishVoice = $speaker.GetInstalledVoices() | Where-Object {
        $_.Enabled -and $_.VoiceInfo.Culture.Name -like 'en-*'
    } | Select-Object -First 1
    if ($englishVoice) { $speaker.SelectVoice($englishVoice.VoiceInfo.Name) }
    $speaker.Rate = -1
    $speaker.Volume = 100
    $speaker.SetOutputToWaveFile($OutputPath)
    $speaker.Speak($text)
} finally {
    $speaker.Dispose()
}
"""
    with tempfile.TemporaryDirectory(prefix="culture-tts-") as temp_dir:
        script_path = Path(temp_dir) / "synthesize.ps1"
        script_path.write_text(script, encoding="utf-8-sig")
        command = [
            "powershell.exe",
            "-NoProfile",
            "-NonInteractive",
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            str(script_path),
            "-TextPath",
            str(text_path),
            "-OutputPath",
            str(output_path),
        ]
        completed = subprocess.run(command, capture_output=True, text=True, encoding="utf-8", errors="replace", timeout=600)
        if completed.returncode != 0:
            detail = completed.stderr.strip() or completed.stdout.strip() or "未知错误"
            raise RuntimeError(f"Windows 语音合成失败：{detail}")


def _generate_local_qr(audio_path: Path, qr_path: Path) -> None:
    import qrcode

    content = audio_path.resolve().as_uri()
    image = qrcode.make(content)
    image.save(qr_path)
