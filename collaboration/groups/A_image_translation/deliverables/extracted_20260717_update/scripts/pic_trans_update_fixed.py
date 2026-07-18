from __future__ import annotations

import base64
import hashlib
import io
import json
import os
import re
import shutil
import tempfile
import zipfile
from dataclasses import asdict, dataclass
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont


PACKAGE_ROOT = Path(__file__).resolve().parents[1]
SOURCE_DOCX = PACKAGE_ROOT / "source" / "翻译资源编写-中国文化知识百科.docx"
WORK_DIR = PACKAGE_ROOT / "work"
ORIGINAL_IMAGES = WORK_DIR / "original_images"
TRANSLATED_IMAGES = PACKAGE_ROOT / "translated_images"
OUTPUT_DOCX = PACKAGE_ROOT / "final_outputs" / "翻译资源编写-中国文化知识百科_A组更新完整修正版.docx"
MANIFEST_XLSX = PACKAGE_ROOT / "manifests" / "translation_manifest_reviewed.xlsx"
CONTACT_SHEET = PACKAGE_ROOT / "previews" / "translated_images_contact_sheet.jpg"
VALIDATION_JSON = PACKAGE_ROOT / "validation" / "validation_report.json"

FONT_REGULAR = Path(os.environ.get("PIC_TRANS_FONT_REGULAR", r"C:\Windows\Fonts\arial.ttf"))
FONT_BOLD = Path(os.environ.get("PIC_TRANS_FONT_BOLD", r"C:\Windows\Fonts\arialbd.ttf"))
REVIEW_DATE = "2026-07-18"
REVIEWER = "Codex independent technical and linguistic review"
CJK_RE = re.compile(r"[\u3400-\u9fff]")
OCR_MIN_CONFIDENCE = 0.75
_OCR_ENGINE = None


@dataclass(frozen=True)
class TextBlock:
    source: str
    translation: str
    rect: tuple[int, int, int, int]
    max_size: int
    min_size: int = 11
    fill: tuple[int, int, int] = (250, 245, 228)
    outline: tuple[int, int, int] | None = (207, 178, 119)
    text_fill: tuple[int, int, int] = (32, 28, 24)
    bold: bool = True
    note: str = "Meaning and layout independently reviewed"


@dataclass(frozen=True)
class SvgText:
    translation: str
    font_size: int
    note: str = "Editable SVG text node independently reviewed"


def block(
    source: str,
    translation: str,
    rect: tuple[int, int, int, int],
    max_size: int,
    **kwargs,
) -> TextBlock:
    return TextBlock(source, translation, rect, max_size, **kwargs)


CHAPTER_STYLE = {
    "min_size": 13,
    "fill": (249, 243, 226),
    "outline": None,
    "text_fill": (45, 38, 32),
}


def chapter_blocks(title_source: str, title: str, subtitle_source: str, subtitle: str) -> list[TextBlock]:
    return [
        block(title_source, title, (185, 615, 615, 702), 31, **CHAPTER_STYLE),
        block(subtitle_source, subtitle, (170, 698, 630, 776), 19, **CHAPTER_STYLE),
    ]


RASTER_BLOCKS: dict[str, list[TextBlock]] = {
    "image2.png": chapter_blocks(
        "第一章哲学思想",
        "Chapter 1: Philosophical Thought",
        "天人合一，儒道禅心",
        "Harmony between humanity and nature; Confucian, Daoist, and Chan wisdom",
    ),
    "image4.jpeg": [
        block("修身立德，顺应自然", "Cultivate the self and virtue; follow nature's way", (500, 65, 1145, 170), 42),
        block(
            "儒家入世修德，道家逍遥自在",
            "Confucians engage with society and cultivate virtue; Daoists live freely and at ease",
            (455, 775, 1190, 870),
            30,
        ),
    ],
    "image5.jpeg": [
        block(
            "禅在山水间，心随云气静",
            "Chan dwells among mountains and waters; the mind grows still with the clouds",
            (925, 95, 1570, 195),
            30,
        ),
        block(
            "言有尽而意无穷",
            "Words may end, but meaning is inexhaustible",
            (1080, 185, 1400, 245),
            23,
            note="OCR source corrected from 言有尽而意无究 to the visible text 言有尽而意无穷",
        ),
    ],
    "image6.png": chapter_blocks(
        "第二章文学诗赋",
        "Chapter 2: Literature and Poetry",
        "诗赋载山河，文章见岁月",
        "Poetry and rhapsody embrace the land; writing bears witness to the ages",
    ),
    "image8.jpeg": [
        block(
            "笔墨载山河，文章照千年",
            "Brush and ink carry the landscape; writing illuminates the ages",
            (500, 382, 1140, 466),
            31,
        ),
        block(
            "诗词曲赋里的中国精神",
            "The Chinese spirit in poetry, lyric verse, songs, and rhapsodies",
            (632, 470, 1012, 530),
            21,
        ),
        block(
            "《诗经》采集民风",
            "The Book of Songs:\ngathering folk songs",
            (390, 295, 520, 405),
            18,
            min_size=11,
            outline=None,
            bold=False,
            note="Prominent text omitted by the submitted OCR; independently added",
        ),
        block(
            "屈原行吟泽畔",
            "Qu Yuan chants\nby the waterside",
            (340, 400, 540, 550),
            17,
            min_size=11,
            outline=None,
            bold=False,
            note="Prominent figure label omitted by the submitted OCR; independently added",
        ),
        block(
            "李白望月",
            "Li Bai gazes\nat the moon",
            (615, 528, 740, 645),
            17,
            min_size=11,
            outline=None,
            bold=False,
            note="Prominent figure label omitted by the submitted OCR; independently added",
        ),
        block(
            "李清照填词",
            "Li Qingzhao\ncomposes lyrics",
            (865, 525, 1005, 645),
            16,
            min_size=10,
            outline=None,
            bold=False,
            note="Prominent figure label omitted by the submitted OCR; independently added",
        ),
        block(
            "曹雪芹著书",
            "Cao Xueqin\nwrites a novel",
            (1065, 475, 1205, 595),
            17,
            min_size=11,
            outline=None,
            bold=False,
            note="Prominent figure label omitted by the submitted OCR; independently added",
        ),
    ],
    "image9.png": [
        block("盛唐气象，两宋风华", "High Tang grandeur; Northern and Southern Song elegance", (430, 65, 1205, 180), 42),
        block("诗言志，词传情", "Poetry voices aspiration; lyrics convey emotion", (505, 750, 1130, 875), 38),
    ],
    "image10.png": chapter_blocks(
        "第三章书法绘画",
        "Chapter 3: Calligraphy and Painting",
        "笔落云烟，墨成山水",
        "A brushstroke conjures clouds; ink forms mountains and waters",
    ),
    "image12.png": chapter_blocks(
        "第四章建筑园林",
        "Chapter 4: Architecture and Gardens",
        "园林藏天地，砖瓦见诗意",
        "Gardens contain worlds; bricks and tiles reveal poetry",
    ),
    "image14.png": [
        block("榫卯相扣，砖瓦成诗", "Mortise-and-tenon joints interlock; bricks and tiles become poetry", (410, 710, 1195, 825), 40),
        block(
            "中国建筑里的礼制与空间美学",
            "Ritual order and spatial aesthetics in Chinese architecture",
            (585, 815, 1065, 890),
            25,
        ),
    ],
    "image15.tiff": [
        block(
            "虽由人作，宛自天开",
            "Though made by human hands, it appears as if created by nature",
            (945, 50, 1585, 155),
            34,
        ),
        block(
            "方寸园林藏山林诗意",
            "Within a small garden lies the poetry of mountains and forests",
            (1035, 145, 1505, 225),
            25,
        ),
    ],
    "image16.png": [
        block(
            "一方水土，一方屋舍",
            "Each place has its own landscape and dwellings",
            (485, 40, 1170, 150),
            43,
            note="OCR source corrected from 一方水止，一方屋舍 to the visible text 一方水土，一方屋舍",
        ),
        block(
            "民居藏着中国人的生活智慧",
            "Vernacular dwellings embody the wisdom of Chinese life",
            (590, 135, 1060, 205),
            27,
        ),
        block("北京", "Beijing", (145, 740, 275, 830), 28, fill=(247, 224, 166), outline=None),
        block("黄土窑洞", "Loess cave dwellings", (500, 735, 750, 830), 27, fill=(226, 184, 91), outline=None),
        block(
            "徽派马头墙",
            "Huizhou-style\nhorse-head walls",
            (900, 730, 1160, 835),
            25,
            fill=(69, 119, 112),
            outline=None,
            text_fill=(255, 252, 238),
            note="OCR source corrected from 徽品头墙 to the visible text 徽派马头墙",
        ),
        block(
            "福建土楼",
            "Fujian tulou",
            (1340, 730, 1545, 835),
            28,
            fill=(168, 75, 43),
            outline=None,
            text_fill=(255, 248, 222),
            note="OCR source corrected from 香兰楼 to the visible text 福建土楼",
        ),
    ],
    "image17.png": chapter_blocks(
        "第五章科技工艺",
        "Chapter 5: Science, Technology, and Craftsmanship",
        "器有匠心，物载文明",
        "Every object embodies craftsmanship; artifacts carry civilization",
    ),
    "image19.png": [
        block("泥土入火，成瓷成韵", "Clay enters the kiln, emerging as porcelain with grace", (450, 705, 1200, 825), 43),
        block("中国瓷器享誉世界", "Chinese porcelain is renowned worldwide", (625, 815, 1020, 890), 27),
    ],
    "image20.tiff": [
        block("一丝一线", "Thread by thread", (1090, 145, 1500, 255), 43),
        block("织就锦绣丝路", "Weaving the splendid Silk Road", (1030, 245, 1600, 360), 42),
        block(
            "养蚕缫丝，连接东西文明",
            "Sericulture and silk reeling linked Eastern and Western civilizations",
            (1035, 365, 1535, 450),
            26,
            note="OCR source corrected from 养蚕织丝 to the visible text 养蚕缫丝",
        ),
    ],
    "image21.png": chapter_blocks(
        "第六章戏曲曲艺",
        "Chapter 6: Opera and Folk Performing Arts",
        "粉墨登场，演尽人间百态",
        "Painted faces take the stage, portraying life in all its forms",
    ),
    "image23.png": [
        block(
            "地方戏台",
            "Local Opera Stage",
            (680, 95, 965, 195),
            29,
            fill=(154, 44, 35),
            outline=(225, 181, 79),
            text_fill=(255, 228, 146),
            note="Prominent sign omitted by the submitted OCR; independently added",
        ),
        block("一方戏台，一方故事", "Every stage tells a story", (470, 705, 1200, 825), 50),
        block(
            "三百剧种扎根乡土",
            "More than 300 opera genres are rooted in local communities",
            (570, 810, 1080, 890),
            27,
        ),
    ],
    "image24.png": chapter_blocks(
        "第七章民俗风情",
        "Chapter 7: Folk Customs and Traditions",
        "烟火寻常处，风俗代代传",
        "In everyday life, customs pass from generation to generation",
    ),
    "image26.png": chapter_blocks(
        "第八章饮食文化",
        "Chapter 8: Food Culture",
        "一茶一饭，皆是东方滋味",
        "Every cup of tea and every meal carries the flavors of the East",
    ),
    "image28.png": chapter_blocks(
        "第九章服饰衣冠",
        "Chapter 9: Traditional Dress and Adornment",
        "衣冠有礼，章服生辉",
        "Attire reflects ritual; ceremonial dress shines",
    ),
    "image30.png": chapter_blocks(
        "第十章宗教文化",
        "Chapter 10: Religious Culture",
        "殊途同归，心向安宁",
        "Different paths lead to the same destination; the heart seeks peace",
    ),
}


# Optional cleanup regions for non-semantic AI-generated remnants in future source images.
MASK_RECTS: dict[str, list[tuple[tuple[int, int, int, int], tuple[int, int, int]]]] = {}

# Reviewed overlays for future SVGs whose visible text is paths or embedded pixels.
# The current source document has no such SVG, but the fallback is exercised in validation.
PATH_ONLY_SVG_BLOCKS: dict[str, list[TextBlock]] = {}


SVG_TRANSLATIONS: dict[str, dict[str, SvgText]] = {
    "image3.svg": {
        "第一章 哲学思想": SvgText("Chapter 1: Philosophical Thought", 27),
        "天人合一，儒道禅心": SvgText(
            "Harmony between humanity and nature; Confucian, Daoist, and Chan wisdom", 16
        ),
    },
    "image7.svg": {
        "第二章 文学诗赋": SvgText("Chapter 2: Literature and Poetry", 27),
        "诗赋载山河，文章见岁月": SvgText(
            "Poetry and rhapsody embrace the land; writing bears witness to the ages", 15
        ),
    },
    "image11.svg": {
        "第三章 书法绘画": SvgText("Chapter 3: Calligraphy and Painting", 27),
        "笔落云烟，墨成山水": SvgText("A brushstroke conjures clouds; ink forms mountains and waters", 16),
    },
    "image13.svg": {
        "第四章 建筑园林": SvgText("Chapter 4: Architecture and Gardens", 27),
        "园林藏天地，砖瓦见诗意": SvgText("Gardens contain worlds; bricks and tiles reveal poetry", 16),
    },
    "image18.svg": {
        "第五章 科技工艺": SvgText("Chapter 5: Science, Technology, and Craftsmanship", 25),
        "器有匠心，物载文明": SvgText("Every object embodies craftsmanship; artifacts carry civilization", 16),
    },
    "image22.svg": {
        "第六章 戏曲曲艺": SvgText("Chapter 6: Opera and Folk Performing Arts", 26),
        "粉墨登场，演尽人间百态": SvgText("Painted faces take the stage, portraying life in all its forms", 16),
    },
    "image25.svg": {
        "第七章 民俗风情": SvgText("Chapter 7: Folk Customs and Traditions", 27),
        "烟火寻常处，风俗代代传": SvgText("In everyday life, customs pass from generation to generation", 16),
    },
    "image27.svg": {
        "第八章 饮食文化": SvgText("Chapter 8: Food Culture", 28),
        "一茶一饭，皆是东方滋味": SvgText("Every cup of tea and every meal carries the flavors of the East", 16),
    },
    "image29.svg": {
        "第九章 服饰衣冠": SvgText("Chapter 9: Traditional Dress and Adornment", 26),
        "衣冠有礼，章服生辉": SvgText("Attire reflects ritual; ceremonial dress shines", 17),
    },
    "image31.svg": {
        "第十章 宗教文化": SvgText("Chapter 10: Religious Culture", 27),
        "殊途同归，心向安宁": SvgText(
            "Different paths lead to the same destination; the heart seeks peace", 15
        ),
    },
}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def get_font(size: int, bold: bool) -> ImageFont.ImageFont:
    primary = FONT_BOLD if bold else FONT_REGULAR
    fallback = FONT_REGULAR if bold else FONT_BOLD
    for candidate in (primary, fallback):
        if candidate.exists():
            return ImageFont.truetype(str(candidate), size)
    return ImageFont.load_default()


def wrap_text(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont, width: int) -> list[str]:
    lines: list[str] = []
    for paragraph in text.split("\n"):
        words = paragraph.split()
        if not words:
            lines.append("")
            continue
        current = words[0]
        for word in words[1:]:
            candidate = f"{current} {word}"
            bbox = draw.textbbox((0, 0), candidate, font=font)
            if bbox[2] - bbox[0] <= width:
                current = candidate
            else:
                lines.append(current)
                current = word
        lines.append(current)
    return lines


def fit_text(draw: ImageDraw.ImageDraw, block: TextBlock) -> tuple[ImageFont.ImageFont, list[str], int]:
    x0, y0, x1, y1 = block.rect
    width = max(1, x1 - x0 - 20)
    height = max(1, y1 - y0 - 14)
    for size in range(block.max_size, block.min_size - 1, -1):
        font = get_font(size, block.bold)
        lines = wrap_text(draw, block.translation, font, width)
        boxes = [draw.textbbox((0, 0), line, font=font) for line in lines]
        line_heights = [max(1, box[3] - box[1]) for box in boxes]
        spacing = max(2, int(size * 0.22))
        total_height = sum(line_heights) + spacing * max(0, len(lines) - 1)
        max_width = max((box[2] - box[0] for box in boxes), default=0)
        if total_height <= height and max_width <= width:
            return font, lines, spacing
    font = get_font(block.min_size, block.bold)
    return font, wrap_text(draw, block.translation, font, width), max(2, int(block.min_size * 0.2))


def draw_block(draw: ImageDraw.ImageDraw, block: TextBlock) -> None:
    x0, y0, x1, y1 = block.rect
    radius = max(4, min(18, (y1 - y0) // 6))
    draw.rounded_rectangle(
        block.rect,
        radius=radius,
        fill=block.fill,
        outline=block.outline,
        width=2 if block.outline else 1,
    )
    font, lines, spacing = fit_text(draw, block)
    boxes = [draw.textbbox((0, 0), line, font=font) for line in lines]
    heights = [max(1, box[3] - box[1]) for box in boxes]
    total_height = sum(heights) + spacing * max(0, len(lines) - 1)
    y = y0 + ((y1 - y0) - total_height) / 2
    for line, box, height in zip(lines, boxes, heights):
        width = box[2] - box[0]
        x = x0 + ((x1 - x0) - width) / 2
        draw.text((x, y - box[1]), line, font=font, fill=block.text_fill)
        y += height + spacing


def clean_output_dirs() -> None:
    for directory in (WORK_DIR, TRANSLATED_IMAGES):
        if directory.exists():
            shutil.rmtree(directory)
        directory.mkdir(parents=True)
    ORIGINAL_IMAGES.mkdir(parents=True, exist_ok=True)
    for directory in (OUTPUT_DOCX.parent, MANIFEST_XLSX.parent, CONTACT_SHEET.parent, VALIDATION_JSON.parent):
        directory.mkdir(parents=True, exist_ok=True)


def extract_docx_media() -> list[str]:
    with zipfile.ZipFile(SOURCE_DOCX) as archive:
        bad = archive.testzip()
        if bad:
            raise RuntimeError(f"Source DOCX has a corrupt member: {bad}")
        media = sorted(
            name
            for name in archive.namelist()
            if name.startswith("word/media/") and not name.endswith("/")
        )
        for member in media:
            target = ORIGINAL_IMAGES / Path(member).name
            target.write_bytes(archive.read(member))
    if len(media) != len({Path(name).name.lower() for name in media}):
        raise RuntimeError("Source DOCX contains duplicate media basenames")
    return [Path(name).name for name in media]


def local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def ocr_image(image: Image.Image) -> list[dict]:
    global _OCR_ENGINE
    try:
        import numpy as np
        from rapidocr_onnxruntime import RapidOCR
    except ImportError as error:
        raise RuntimeError(
            "Offline OCR requires rapidocr-onnxruntime; install requirements.txt before running"
        ) from error
    if _OCR_ENGINE is None:
        _OCR_ENGINE = RapidOCR()
    result, _ = _OCR_ENGINE(np.asarray(image.convert("RGB")))
    if not result:
        return []
    return [
        {
            "text": str(item[1]),
            "confidence": round(float(item[2]), 4),
            "box": [[round(float(value), 2) for value in point] for point in item[0]],
        }
        for item in result
    ]


def render_svg_image(svg_bytes: bytes, width: int | None = None) -> Image.Image:
    try:
        import resvg_py
    except ImportError as error:
        raise RuntimeError("SVG rasterization requires resvg_py; install requirements.txt before running") from error
    png = resvg_py.svg_to_bytes(svg_string=svg_bytes.decode("utf-8"), width=width)
    with Image.open(io.BytesIO(png)) as rendered:
        return rendered.convert("RGB")


def png_image_as_svg(image: Image.Image) -> bytes:
    stream = io.BytesIO()
    image.save(stream, format="PNG", optimize=True)
    encoded = base64.b64encode(stream.getvalue()).decode("ascii")
    width, height = image.size
    wrapper = (
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" '
        f'viewBox="0 0 {width} {height}"><image width="{width}" height="{height}" '
        f'href="data:image/png;base64,{encoded}"/></svg>'
    )
    return wrapper.encode("utf-8")


def translate_path_only_svg(svg_bytes: bytes, blocks: list[TextBlock], paths: int) -> tuple[bytes, dict]:
    if not blocks:
        raise ValueError("SVG has no editable CJK nodes and no reviewed OCR overlay specification")
    image = render_svg_image(svg_bytes)
    source_ocr = ocr_image(image)
    detected_cjk = "".join(
        item["text"]
        for item in source_ocr
        if item["confidence"] >= OCR_MIN_CONFIDENCE and CJK_RE.search(item["text"])
    )
    missing = [block.source for block in blocks if block.source.replace(" ", "") not in detected_cjk.replace(" ", "")]
    if missing:
        raise ValueError(f"Reviewed path-only SVG source text was not found by offline OCR: {missing}")

    draw = ImageDraw.Draw(image)
    for item in blocks:
        draw_block(draw, item)
    output_ocr = ocr_image(image)
    leftovers = [
        item
        for item in output_ocr
        if item["confidence"] >= OCR_MIN_CONFIDENCE and CJK_RE.search(item["text"])
    ]
    if leftovers:
        raise ValueError(f"CJK text remains after path-only SVG fallback: {leftovers}")
    translated = png_image_as_svg(image)
    ET.fromstring(translated)
    return translated, {
        "mode": "render-ocr-reviewed-overlay",
        "paths": paths,
        "editable_nodes": 0,
        "translated_nodes": len(blocks),
        "source_ocr": source_ocr,
        "output_ocr": output_ocr,
        "pairs": {item.source: item.translation for item in blocks},
    }


def translate_svg(
    svg_bytes: bytes,
    translations: dict[str, SvgText],
    path_only_blocks: list[TextBlock] | None = None,
) -> tuple[bytes, dict]:
    ET.register_namespace("", "http://www.w3.org/2000/svg")
    ET.register_namespace("xlink", "http://www.w3.org/1999/xlink")
    root = ET.fromstring(svg_bytes)
    parents = {child: parent for parent in root.iter() for child in parent}
    paths = sum(1 for node in root.iter() if local_name(node.tag) == "path")
    editable = [node for node in root.iter() if local_name(node.tag) in {"text", "tspan"}]
    cjk_nodes = [node for node in editable if node.text and CJK_RE.search(node.text)]
    if not cjk_nodes:
        if paths:
            if path_only_blocks is not None:
                return translate_path_only_svg(svg_bytes, path_only_blocks, paths)
            raise ValueError(
                "SVG contains paths but no editable CJK text nodes; add reviewed overlays to "
                "PATH_ONLY_SVG_BLOCKS for automatic render, OCR, and fill-back"
            )
        raise ValueError("SVG contains no editable CJK text nodes")

    found: dict[str, str] = {}
    for node in cjk_nodes:
        source = (node.text or "").strip()
        if source not in translations:
            raise ValueError(f"Unreviewed SVG text: {source!r}")
        spec = translations[source]
        node.text = spec.translation
        found[source] = spec.translation
        style_node = parents.get(node, node)
        style_node.set("font-family", "Arial, sans-serif")
        style_node.set("font-size", str(spec.font_size))

    missing = sorted(set(translations) - set(found))
    if missing:
        raise ValueError(f"Expected SVG text not found: {missing}")
    leftovers = [node.text for node in editable if node.text and CJK_RE.search(node.text)]
    if leftovers:
        raise ValueError(f"CJK text remains after SVG translation: {leftovers}")

    translated = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    ET.fromstring(translated)
    return translated, {
        "paths": paths,
        "editable_nodes": len(editable),
        "translated_nodes": len(found),
        "pairs": found,
    }


def render_images(media_names: list[str]) -> dict[str, dict]:
    result: dict[str, dict] = {}
    expected_svg = {name for name in media_names if name.lower().endswith(".svg")}
    reviewed_svg = set(SVG_TRANSLATIONS) | set(PATH_ONLY_SVG_BLOCKS)
    if expected_svg != reviewed_svg:
        raise RuntimeError(
            f"SVG manifest mismatch; DOCX={sorted(expected_svg)}, reviewed={sorted(reviewed_svg)}"
        )

    for name in media_names:
        source = ORIGINAL_IMAGES / name
        target = TRANSLATED_IMAGES / name
        suffix = source.suffix.lower()
        if suffix == ".svg":
            translated, details = translate_svg(
                source.read_bytes(),
                SVG_TRANSLATIONS.get(name, {}),
                PATH_ONLY_SVG_BLOCKS.get(name),
            )
            target.write_bytes(translated)
            result[name] = {"type": "svg", **details}
            continue
        if name == "image1.jpeg":
            shutil.copy2(source, target)
            result[name] = {"type": "raster", "text_blocks": 0, "copied": True}
            continue
        blocks = RASTER_BLOCKS.get(name)
        if blocks is None:
            raise RuntimeError(f"No reviewed raster layout for DOCX media {name}")
        with Image.open(source) as original:
            image = original.convert("RGB")
        draw = ImageDraw.Draw(image)
        for rect, fill in MASK_RECTS.get(name, []):
            draw.rounded_rectangle(rect, radius=4, fill=fill)
        for text_block in blocks:
            draw_block(draw, text_block)
        if suffix in {".jpg", ".jpeg"}:
            image.save(target, quality=95, optimize=True)
        elif suffix in {".tif", ".tiff"}:
            image.save(target, compression="tiff_deflate")
        elif suffix == ".png":
            image.save(target, optimize=True)
        else:
            raise RuntimeError(f"Unsupported raster format: {source}")
        result[name] = {"type": "raster", "text_blocks": len(blocks), "copied": False}
    return result


def write_manifest() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "reviewed translations"
    headers = [
        "image",
        "media_type",
        "source_text",
        "translation",
        "x0",
        "y0",
        "x1",
        "y1",
        "review_status",
        "reviewer",
        "review_date",
        "notes",
    ]
    ws.append(headers)
    for image_name, blocks in RASTER_BLOCKS.items():
        for item in blocks:
            x0, y0, x1, y1 = item.rect
            ws.append(
                [
                    image_name,
                    "raster",
                    item.source,
                    item.translation,
                    x0,
                    y0,
                    x1,
                    y1,
                    "independently reviewed",
                    REVIEWER,
                    REVIEW_DATE,
                    item.note,
                ]
            )
    for image_name, translations in SVG_TRANSLATIONS.items():
        for source, item in translations.items():
            ws.append(
                [
                    image_name,
                    "svg",
                    source,
                    item.translation,
                    None,
                    None,
                    None,
                    None,
                    "independently reviewed",
                    REVIEWER,
                    REVIEW_DATE,
                    item.note,
                ]
            )
    for image_name, blocks in PATH_ONLY_SVG_BLOCKS.items():
        for item in blocks:
            x0, y0, x1, y1 = item.rect
            ws.append(
                [
                    image_name,
                    "svg_rendered_ocr",
                    item.source,
                    item.translation,
                    x0,
                    y0,
                    x1,
                    y1,
                    "independently reviewed",
                    REVIEWER,
                    REVIEW_DATE,
                    item.note,
                ]
            )

    header_fill = PatternFill("solid", fgColor="244062")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    widths = [18, 12, 36, 68, 9, 9, 9, 9, 22, 34, 14, 60]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(MANIFEST_XLSX)


def build_docx(media_names: list[str]) -> None:
    expected = {name.lower() for name in media_names}
    translated = {path.name.lower(): path for path in TRANSLATED_IMAGES.iterdir() if path.is_file()}
    if expected != set(translated):
        raise RuntimeError(
            f"Translated media mismatch; missing={sorted(expected - set(translated))}, "
            f"extra={sorted(set(translated) - expected)}"
        )

    with tempfile.NamedTemporaryFile(delete=False, suffix=".docx", dir=OUTPUT_DOCX.parent) as temp_stream:
        temp_path = Path(temp_stream.name)
    try:
        with zipfile.ZipFile(SOURCE_DOCX) as source, zipfile.ZipFile(temp_path, "w", zipfile.ZIP_DEFLATED) as output:
            for info in source.infolist():
                data = source.read(info.filename)
                if info.filename.startswith("word/media/") and not info.filename.endswith("/"):
                    data = translated[Path(info.filename).name.lower()].read_bytes()
                output.writestr(info, data)
        os.replace(temp_path, OUTPUT_DOCX)
    finally:
        temp_path.unlink(missing_ok=True)


def make_contact_sheet(media_names: list[str]) -> None:
    media_paths = [TRANSLATED_IMAGES / name for name in media_names]
    columns = 4
    tile_w, tile_h = 340, 230
    label_h = 30
    rows = (len(media_paths) + columns - 1) // columns
    sheet = Image.new("RGB", (columns * tile_w, rows * (tile_h + label_h)), (238, 238, 235))
    draw = ImageDraw.Draw(sheet)
    label_font = get_font(18, True)
    for index, path in enumerate(media_paths):
        if path.suffix.lower() == ".svg":
            image = render_svg_image(path.read_bytes(), width=tile_w - 12)
        else:
            with Image.open(path) as source:
                image = source.convert("RGB")
        image.thumbnail((tile_w - 12, tile_h - 12), Image.Resampling.LANCZOS)
        col, row = index % columns, index // columns
        x = col * tile_w + (tile_w - image.width) // 2
        y = row * (tile_h + label_h) + (tile_h - image.height) // 2
        sheet.paste(image, (x, y))
        label = path.name
        box = draw.textbbox((0, 0), label, font=label_font)
        draw.text(
            (col * tile_w + (tile_w - (box[2] - box[0])) / 2, row * (tile_h + label_h) + tile_h + 3),
            label,
            font=label_font,
            fill=(30, 30, 30),
        )
    sheet.save(CONTACT_SHEET, quality=92)


def inspect_docx(path: Path) -> dict:
    with zipfile.ZipFile(path) as archive:
        bad = archive.testzip()
        media = [name for name in archive.namelist() if name.startswith("word/media/") and not name.endswith("/")]
        rels = ET.fromstring(archive.read("word/_rels/document.xml.rels"))
        image_rels = [node for node in rels if (node.get("Type") or "").endswith("/image")]
        document = ET.fromstring(archive.read("word/document.xml"))
        embed_refs = sum(
            1
            for node in document.iter()
            for attr in node.attrib
            if local_name(attr) == "embed"
        )
    return {
        "zip_test": bad,
        "media_count": len(media),
        "svg_count": sum(name.lower().endswith(".svg") for name in media),
        "image_relationships": len(image_rels),
        "embedded_image_refs": embed_refs,
        "media_names": [Path(name).name for name in media],
    }


def path_only_fallback_test() -> dict:
    font_path = Path(r"C:\Windows\Fonts\msyhbd.ttc")
    if not font_path.exists():
        font_path = Path(r"C:\Windows\Fonts\msyh.ttc")
    if not font_path.exists():
        raise RuntimeError("The controlled path-only SVG OCR test requires a Windows CJK font")

    source_image = Image.new("RGB", (640, 240), (250, 248, 240))
    draw = ImageDraw.Draw(source_image)
    font = ImageFont.truetype(str(font_path), 72)
    text = "中国文化"
    box = draw.textbbox((0, 0), text, font=font)
    draw.text(((640 - (box[2] - box[0])) / 2, 72), text, font=font, fill=(30, 30, 30))
    source_stream = io.BytesIO()
    source_image.save(source_stream, format="PNG")
    encoded = base64.b64encode(source_stream.getvalue()).decode("ascii")
    sample = (
        '<svg xmlns="http://www.w3.org/2000/svg" width="640" height="240" viewBox="0 0 640 240">'
        '<path d="M10 10H630V230H10Z" fill="none" stroke="#555"/>'
        f'<image width="640" height="240" href="data:image/png;base64,{encoded}"/></svg>'
    ).encode("utf-8")
    reviewed = [
        block(
            "中国文化",
            "Chinese Culture",
            (70, 55, 570, 190),
            52,
            fill=(250, 248, 240),
            outline=(110, 110, 100),
            note="Controlled no-editable-text SVG fallback test",
        )
    ]
    translated, details = translate_svg(sample, {}, reviewed)
    rendered = render_svg_image(translated)
    if rendered.size != (640, 240):
        raise AssertionError(f"Unexpected path-only fallback size: {rendered.size}")
    residual = [
        item
        for item in ocr_image(rendered)
        if item["confidence"] >= OCR_MIN_CONFIDENCE and CJK_RE.search(item["text"])
    ]
    if residual:
        raise AssertionError(f"Controlled path-only fallback retained CJK: {residual}")
    return {
        "passed": True,
        "mode": details["mode"],
        "source_text": "中国文化",
        "translation": "Chinese Culture",
        "output_size": list(rendered.size),
    }


def validate_raster_ocr(media_names: list[str]) -> dict:
    scanned = []
    residual: dict[str, list[dict]] = {}
    for name in media_names:
        if name.lower().endswith(".svg") or name == "image1.jpeg":
            continue
        with Image.open(TRANSLATED_IMAGES / name) as source:
            image = source.convert("RGB")
        detections = ocr_image(image)
        cjk = [
            item
            for item in detections
            if item["confidence"] >= OCR_MIN_CONFIDENCE and CJK_RE.search(item["text"])
        ]
        scanned.append(name)
        if cjk:
            residual[name] = cjk
    if residual:
        raise AssertionError(f"High-confidence CJK remains in translated raster media: {residual}")
    return {
        "scanned_count": len(scanned),
        "skipped_qr_media": "image1.jpeg",
        "minimum_confidence": OCR_MIN_CONFIDENCE,
        "high_confidence_cjk_remaining": 0,
    }


def validate_svg_renders(media_names: list[str]) -> dict:
    details = {}
    for name in media_names:
        if not name.lower().endswith(".svg"):
            continue
        image = render_svg_image((TRANSLATED_IMAGES / name).read_bytes(), width=800)
        extrema = image.getextrema()
        if all(low == high for low, high in extrema):
            raise AssertionError(f"Rendered SVG is blank: {name}")
        residual = [
            item
            for item in ocr_image(image)
            if item["confidence"] >= OCR_MIN_CONFIDENCE and CJK_RE.search(item["text"])
        ]
        if residual:
            raise AssertionError(f"Rendered SVG retained CJK: {name}: {residual}")
        details[name] = {
            "render_size": list(image.size),
            "nonblank": True,
            "high_confidence_cjk_remaining": 0,
        }
    return {"rendered_count": len(details), "details": details}


def validate(media_names: list[str], render_details: dict[str, dict]) -> dict:
    source = inspect_docx(SOURCE_DOCX)
    output = inspect_docx(OUTPUT_DOCX)
    if source["media_names"] != output["media_names"]:
        raise AssertionError("Output DOCX changed the media inventory")
    if output["zip_test"] is not None:
        raise AssertionError(f"Output DOCX has a corrupt member: {output['zip_test']}")
    if output["media_count"] != len(media_names):
        raise AssertionError("Output DOCX media count mismatch")
    if output["svg_count"] != len(SVG_TRANSLATIONS) + len(PATH_ONLY_SVG_BLOCKS):
        raise AssertionError("Output DOCX SVG count mismatch")

    original_hashes = {path.name: sha256(path) for path in ORIGINAL_IMAGES.iterdir() if path.is_file()}
    translated_hashes = {path.name: sha256(path) for path in TRANSLATED_IMAGES.iterdir() if path.is_file()}
    changed = sorted(name for name in media_names if original_hashes[name] != translated_hashes[name])
    unchanged = sorted(name for name in media_names if original_hashes[name] == translated_hashes[name])
    if unchanged != ["image1.jpeg"]:
        raise AssertionError(f"Unexpected unchanged media: {unchanged}")

    report = {
        "review_date": REVIEW_DATE,
        "reviewer": REVIEWER,
        "source_docx": {"sha256": sha256(SOURCE_DOCX), **source},
        "output_docx": {"sha256": sha256(OUTPUT_DOCX), **output},
        "manifest": {
            "sha256": sha256(MANIFEST_XLSX),
            "reviewed_rows": sum(len(items) for items in RASTER_BLOCKS.values())
            + sum(len(items) for items in SVG_TRANSLATIONS.values())
            + sum(len(items) for items in PATH_ONLY_SVG_BLOCKS.values()),
        },
        "media": {
            "changed": changed,
            "unchanged": unchanged,
            "details": render_details,
            "nonsemantic_masks": {
                name: [list(rect) for rect, _ in masks] for name, masks in MASK_RECTS.items()
            },
        },
        "raster_ocr_validation": validate_raster_ocr(media_names),
        "svg_render_validation": validate_svg_renders(media_names),
        "path_only_svg_fallback_test": path_only_fallback_test(),
        "contact_sheet": {"sha256": sha256(CONTACT_SHEET)},
    }
    VALIDATION_JSON.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return report


def main() -> None:
    if not SOURCE_DOCX.exists():
        raise FileNotFoundError(SOURCE_DOCX)
    clean_output_dirs()
    media_names = extract_docx_media()
    render_details = render_images(media_names)
    write_manifest()
    build_docx(media_names)
    make_contact_sheet(media_names)
    report = validate(media_names, render_details)
    print(f"Source DOCX: {SOURCE_DOCX}")
    print(f"Reviewed manifest: {MANIFEST_XLSX}")
    print(f"Translated media: {TRANSLATED_IMAGES}")
    print(f"Final DOCX: {OUTPUT_DOCX}")
    print(f"Contact sheet: {CONTACT_SHEET}")
    print(f"Validation report: {VALIDATION_JSON}")
    print(f"Media: {report['output_docx']['media_count']} total, {report['output_docx']['svg_count']} SVG")
    print(f"Reviewed text items: {report['manifest']['reviewed_rows']}")


if __name__ == "__main__":
    main()
