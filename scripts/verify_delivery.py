from __future__ import annotations

import hashlib
import json
import re
import sys
import zipfile
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass(frozen=True)
class CheckResult:
    name: str
    passed: bool
    detail: str


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def main() -> int:
    for stream in (sys.stdout, sys.stderr):
        reconfigure = getattr(stream, "reconfigure", None)
        if callable(reconfigure):
            reconfigure(encoding="utf-8", errors="replace")

    root = Path(__file__).resolve().parents[1]
    checks: list[CheckResult] = []

    required_paths = {
        "Qt 主程序": root / "src/agent_gui_starter/app.py",
        "OpenAI 智能体": root / "src/agent_gui_starter/agent.py",
        "Coze 工作流客户端": root / "src/agent_gui_starter/coze.py",
        "统一生产处理层": root / "src/agent_gui_starter/production.py",
        "总整合验收记录": root
        / "collaboration/integration/final_outputs/INTEGRATION_ACCEPTANCE_20260718.md",
        "图文更新修正版压缩包": root
        / "collaboration/groups/A_image_translation/deliverables/archives/pic_trans_update_fixed_20260717.zip",
        "图文更新版最终 DOCX": root
        / "collaboration/groups/A_image_translation/deliverables/extracted_20260717_update/final_outputs/翻译资源编写-中国文化知识百科_A组更新完整修正版.docx",
        "图文 17 页渲染总览": root
        / "collaboration/groups/A_image_translation/deliverables/extracted_20260717_update/previews/final_docx_pages_contact_sheet.jpg",
        "图文机器可读校验": root
        / "collaboration/groups/A_image_translation/deliverables/extracted_20260717_update/validation/validation_report.json",
        "文化术语工作流校验": root
        / "collaboration/groups/B_terms_style/deliverables/workflow_translation_draft_3045/workflow_validation.json",
        "音频测试样例": root
        / "collaboration/groups/C_text_audio_translation/deliverables/audio_video_workflow/revised_20260717/supplement/测试音频.mp3",
        "音频终版成品": root
        / "collaboration/groups/C_text_audio_translation/deliverables/audio_video_workflow/revised_20260717/supplement/模式二生成总音频.mp3",
        "音频终版审校表": root
        / "collaboration/groups/C_text_audio_translation/deliverables/audio_video_workflow/revised_20260717/supplement/模式二生成终版表格.xlsx",
        "音频播放二维码": root
        / "collaboration/groups/C_text_audio_translation/deliverables/audio_video_workflow/revised_20260717/supplement/模式二生成二维码.png",
        "统一成品使用说明": root
        / "collaboration/integration/final_outputs/ready_to_use/README.md",
        "统一成品资源索引": root
        / "collaboration/integration/final_outputs/ready_to_use/05_资源索引/统一交付资源索引.xlsx",
        "统一成品最终验收": root
        / "collaboration/integration/final_outputs/ready_to_use/05_资源索引/最终验收.json",
    }
    for name, path in required_paths.items():
        checks.append(CheckResult(name, path.is_file() and path.stat().st_size > 0, str(path.relative_to(root))))

    terms_path = root / "collaboration/shared/terminology/zhonghua_culture_terms.normalized.json"
    try:
        terms = json.loads(terms_path.read_text(encoding="utf-8"))
        term_count = len(terms) if isinstance(terms, list) else 0
        checks.append(CheckResult("共享术语库", term_count == 212, f"{term_count} 条"))
    except (OSError, json.JSONDecodeError) as exc:
        checks.append(CheckResult("共享术语库", False, str(exc)))

    corpus_path = root / "collaboration/shared/corpus/official_culture_terms_20260718.json"
    try:
        corpus = json.loads(corpus_path.read_text(encoding="utf-8"))
        source_links = [str(item.get("来源链接", "")) for item in corpus if isinstance(item, dict)]
        valid = len(corpus) == 40 and all(link.startswith("https://") for link in source_links)
        checks.append(CheckResult("官方文化补充语料", valid, f"{len(corpus)} 条 / {len(source_links)} 条来源链接"))
    except (OSError, json.JSONDecodeError) as exc:
        checks.append(CheckResult("官方文化补充语料", False, str(exc)))

    a_manifest = root / (
        "collaboration/groups/A_image_translation/deliverables/extracted_20260717_update/"
        "manifests/translation_manifest_reviewed.xlsx"
    )
    try:
        workbook = load_workbook(a_manifest, read_only=True, data_only=True)
        sheet = workbook["reviewed translations"]
        rows = [row for row in sheet.iter_rows(min_row=2, values_only=True) if any(value not in (None, "") for value in row)]
        headers = tuple(cell.value for cell in sheet[1])
        required_headers = {"image", "media_type", "source_text", "translation", "review_status", "reviewer", "review_date"}
        review_index = headers.index("review_status")
        complete_reviews = sum(1 for row in rows if row[review_index] not in (None, ""))
        workbook.close()
        valid = len(rows) == 71 and required_headers.issubset(headers) and complete_reviews == 71
        checks.append(CheckResult("图文更新版翻译清单", valid, f"{len(rows)} 条数据 / {complete_reviews} 条复核状态"))
    except Exception as exc:  # Workbook parser errors must fail delivery verification.
        checks.append(CheckResult("图文更新版翻译清单", False, str(exc)))

    a_validation_path = required_paths["图文机器可读校验"]
    a_docx_path = required_paths["图文更新版最终 DOCX"]
    try:
        validation = json.loads(a_validation_path.read_text(encoding="utf-8"))
        rendered_pages = a_validation_path.parent / "rendered_pages"
        with zipfile.ZipFile(a_docx_path) as document:
            docx_ok = document.testzip() is None
        structure_ok = (
            validation.get("manifest", {}).get("reviewed_rows") == 71
            and validation.get("output_docx", {}).get("sha256") == _sha256(a_docx_path)
            and validation.get("output_docx", {}).get("media_count") == 31
            and validation.get("output_docx", {}).get("svg_count") == 10
            and validation.get("word_render", {}).get("export_succeeded") is True
            and validation.get("word_render", {}).get("page_count") == 17
            and validation.get("word_render", {}).get("rendered_pages_count") == 17
            and len(list(rendered_pages.glob("page_*.png"))) == 17
            and validation.get("path_only_svg_fallback_test", {}).get("passed") is True
            and docx_ok
        )
        checks.append(CheckResult("图文 DOCX 与渲染结构", structure_ok, "31 媒体 / 10 SVG / 17 页 / 哈希一致"))
    except (OSError, json.JSONDecodeError, zipfile.BadZipFile) as exc:
        checks.append(CheckResult("图文 DOCX 与渲染结构", False, str(exc)))

    a_archive_path = required_paths["图文更新修正版压缩包"]
    try:
        credential_pattern = re.compile(
            rb"(?i)(api[_ -]?key|secret|token|password)\s*[=:]\s*[\"']?[A-Za-z0-9_-]{12,}"
        )
        text_extensions = {".py", ".md", ".txt", ".json", ".yml", ".yaml", ".env", ".ini", ".cfg"}
        with zipfile.ZipFile(a_archive_path) as archive:
            names = archive.namelist()
            unsafe_names = [
                name for name in names if name.startswith(("/", "\\")) or ".." in Path(name).parts
            ]
            translated_images = [
                name for name in names if "/translated_images/" in name and not name.endswith("/")
            ]
            credential_files = [
                name
                for name in names
                if Path(name).suffix.lower() in text_extensions
                and credential_pattern.search(archive.read(name))
            ]
            archive_ok = archive.testzip() is None
        valid = archive_ok and not unsafe_names and not credential_files and len(translated_images) == 31
        checks.append(CheckResult("图文压缩包安全与完整性", valid, f"31 个译后媒体 / 凭据命中 {len(credential_files)}"))
    except (OSError, zipfile.BadZipFile) as exc:
        checks.append(CheckResult("图文压缩包安全与完整性", False, str(exc)))

    validation_path = required_paths["文化术语工作流校验"]
    try:
        validation = json.loads(validation_path.read_text(encoding="utf-8"))
        code_nodes = validation.get("code_node_tests", [])
        valid = (
            validation.get("workflow_id") == "7661678571702747178"
            and validation.get("node_count") == 18
            and validation.get("edge_count") == 28
            and validation.get("graph_valid") is True
            and validation.get("all_code_nodes_pass") is True
            and len(code_nodes) == 3
        )
        checks.append(CheckResult("文化术语工作流结构", valid, "18 节点 / 28 边 / 3 个代码节点"))
    except (OSError, json.JSONDecodeError) as exc:
        checks.append(CheckResult("文化术语工作流结构", False, str(exc)))

    c_cases = root / "collaboration/groups/C_text_audio_translation/deliverables/docx_translation/revised_20260717/test_cases"
    case_count = len([path for path in c_cases.iterdir() if path.is_dir()]) if c_cases.is_dir() else 0
    checks.append(CheckResult("DOCX 测试样例", case_count == 5, f"{case_count} 套"))

    ready_root = root / "collaboration/integration/final_outputs/ready_to_use"
    acceptance_path = ready_root / "05_资源索引/最终验收.json"
    try:
        acceptance = json.loads(acceptance_path.read_text(encoding="utf-8"))
        ready_files = [path for path in ready_root.rglob("*") if path.is_file()]
        required_channels = {"01_图文翻译", "02_术语与风格", "03_DOCX翻译", "04_音视频翻译"}
        actual_channels = {path.relative_to(ready_root).parts[0] for path in ready_files}
        valid = (
            acceptance.get("passed") is True
            and acceptance.get("catalog_has_sha256") is True
            and required_channels.issubset(actual_channels)
            and len(ready_files) >= 80
            and any(path.suffix.lower() == ".wav" and path.stat().st_size > 1024 for path in ready_files)
        )
        checks.append(CheckResult("统一直接使用成品区", valid, f"{len(ready_files)} 个文件 / 4 条生产通道"))
    except (OSError, json.JSONDecodeError) as exc:
        checks.append(CheckResult("统一直接使用成品区", False, str(exc)))

    failures = [check for check in checks if not check.passed]
    for check in checks:
        state = "PASS" if check.passed else "FAIL"
        print(f"[{state}] {check.name}: {check.detail}")

    print(f"delivery_checks={len(checks)} failures={len(failures)}")
    return 1 if failures else 0


if __name__ == "__main__":
    sys.exit(main())
