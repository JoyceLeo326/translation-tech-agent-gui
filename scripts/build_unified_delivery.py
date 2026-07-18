from __future__ import annotations

import hashlib
import json
import shutil
import zipfile
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
COLLABORATION = ROOT / "collaboration"
TARGET = COLLABORATION / "integration/final_outputs/ready_to_use"


def main() -> int:
    _reset_target()
    image_source = COLLABORATION / "groups/A_image_translation/deliverables/extracted_20260717_update"
    terms_source = COLLABORATION / "shared/terminology"
    corpus_source = COLLABORATION / "shared/corpus"
    style_source = COLLABORATION / "groups/B_terms_style/prompts"
    docx_source = (
        COLLABORATION
        / "groups/C_text_audio_translation/deliverables/docx_translation/revised_20260717"
    )
    audio_source = (
        COLLABORATION
        / "groups/C_text_audio_translation/deliverables/audio_video_workflow/revised_20260717"
    )
    generated = COLLABORATION / "integration/final_outputs/generated"

    copy_plan = [
        (
            image_source / "final_outputs/翻译资源编写-中国文化知识百科_A组更新完整修正版.docx",
            TARGET / "01_图文翻译/中国文化知识百科_图文英文回填终版.docx",
        ),
        (
            image_source / "manifests/translation_manifest_reviewed.xlsx",
            TARGET / "01_图文翻译/图中文字_71条人工审校清单.xlsx",
        ),
        (
            image_source / "previews/final_docx_pages_contact_sheet.jpg",
            TARGET / "01_图文翻译/17页最终文档预览.jpg",
        ),
        (
            image_source / "previews/translated_images_contact_sheet.jpg",
            TARGET / "01_图文翻译/图文回填素材预览.jpg",
        ),
        (
            image_source / "validation/validation_report.json",
            TARGET / "01_图文翻译/图文回填结构与渲染验收.json",
        ),
        (
            terms_source / "中华文化术语对照表.xlsx",
            TARGET / "02_术语与风格/基础文化术语库_212条/中华文化术语对照表.xlsx",
        ),
        (
            terms_source / "zhonghua_culture_terms.normalized.csv",
            TARGET / "02_术语与风格/基础文化术语库_212条/中华文化术语.normalized.csv",
        ),
        (
            terms_source / "zhonghua_culture_terms.normalized.json",
            TARGET / "02_术语与风格/基础文化术语库_212条/中华文化术语.normalized.json",
        ),
        (corpus_source, TARGET / "02_术语与风格/官方文化补充语料"),
        (style_source, TARGET / "02_术语与风格/儿童文学风格提示词"),
        (docx_source / "test_cases", TARGET / "03_DOCX翻译/五套完整测试样例"),
        (audio_source / "supplement", TARGET / "04_音视频翻译/完整音频测试与成品"),
    ]
    for source, destination in copy_plan:
        _copy(source, destination)
    _extract_docx_media(
        image_source / "final_outputs/翻译资源编写-中国文化知识百科_A组更新完整修正版.docx",
        TARGET / "01_图文翻译/已翻译图片资源",
    )

    docx_output = TARGET / "03_DOCX翻译/本机重新验收产出"
    for sample_index in range(1, 6):
        for suffix in (".docx", ".回填报告.json"):
            candidates = sorted(
                generated.glob(f"DOCX样例_测试{sample_index}_*{suffix}"),
                key=lambda path: path.stat().st_mtime,
            )
            if candidates:
                _copy(candidates[-1], docx_output / candidates[-1].name)
    acceptance_reports = sorted(
        generated.glob("DOCX_五套样例端到端验收_*.json"),
        key=lambda path: path.stat().st_mtime,
    )
    if acceptance_reports:
        _copy(acceptance_reports[-1], docx_output / acceptance_reports[-1].name)
    for pattern in ("*英文配音_*.wav", "*英文配音_*.txt", "*英文配音_*.二维码.png"):
        candidates = sorted(generated.glob(pattern), key=lambda path: path.stat().st_mtime)
        if candidates:
            _copy(candidates[-1], TARGET / "04_音视频翻译/本机重新生成产出" / candidates[-1].name)

    records = _build_catalog()
    _write_catalog(records)
    _write_readme(records)
    _write_acceptance(records)
    print(f"Unified delivery ready: {TARGET}")
    print(f"Files: {len(records)}")
    return 0


def _reset_target() -> None:
    resolved_root = COLLABORATION.resolve()
    resolved_target = TARGET.resolve()
    if resolved_root not in resolved_target.parents:
        raise RuntimeError(f"Refusing to rebuild unexpected path: {resolved_target}")
    if TARGET.exists():
        shutil.rmtree(TARGET)
    TARGET.mkdir(parents=True)


def _copy(source: Path, destination: Path) -> None:
    if not source.exists():
        raise FileNotFoundError(f"Required delivery asset is missing: {source}")
    destination.parent.mkdir(parents=True, exist_ok=True)
    if source.is_dir():
        shutil.copytree(source, destination, dirs_exist_ok=True)
    else:
        shutil.copy2(source, destination)


def _extract_docx_media(source_docx: Path, destination: Path) -> None:
    destination.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(source_docx) as archive:
        bad_member = archive.testzip()
        if bad_member:
            raise RuntimeError(f"Corrupt DOCX member: {bad_member}")
        for member in archive.namelist():
            if member.startswith("word/media/") and not member.endswith("/"):
                (destination / Path(member).name).write_bytes(archive.read(member))


def _build_catalog() -> list[dict[str, object]]:
    records: list[dict[str, object]] = []
    for path in sorted(item for item in TARGET.rglob("*") if item.is_file()):
        if path.name in {"README.md", "统一交付资源索引.xlsx", "统一交付资源索引.json", "最终验收.json"}:
            continue
        records.append(
            {
                "任务通道": path.relative_to(TARGET).parts[0],
                "文件名": path.name,
                "格式": path.suffix.lstrip(".").upper() or "FILE",
                "大小Bytes": path.stat().st_size,
                "SHA256": _sha256(path),
                "相对路径": path.relative_to(TARGET).as_posix(),
            }
        )
    return records


def _write_catalog(records: list[dict[str, object]]) -> None:
    json_path = TARGET / "05_资源索引/统一交付资源索引.json"
    json_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "统一交付资源"
    headers = ["任务通道", "文件名", "格式", "大小Bytes", "SHA256", "相对路径"]
    sheet.append(headers)
    for record in records:
        sheet.append([record[header] for header in headers])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [20, 56, 12, 16, 68, 100]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="245FCC")
        cell.alignment = Alignment(horizontal="center")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    workbook.save(TARGET / "05_资源索引/统一交付资源索引.xlsx")


def _write_readme(records: list[dict[str, object]]) -> None:
    counts: dict[str, int] = {}
    for record in records:
        channel = str(record["任务通道"])
        counts[channel] = counts.get(channel, 0) + 1
    lines = [
        "# 文澜 · 中国文化多模态外译统一交付",
        "",
        "这里是最终用户入口，内容按实际任务流程归类，不按原协作分组区分。",
        "",
        "## 使用顺序",
        "",
        "1. 在桌面程序的“开始”页拖入图片、Word、音频或视频。",
        "2. Word 和音视频会生成 Excel 译文确认表，自动翻译后由人工确认。",
        "3. Word 根据确认表生成保留版式的英文文档，音视频生成英文配音。",
        "4. 所有新成品在“我的成品”查看；本目录保存完整初始成品和测试样例。",
        "",
        "## 已整理内容",
        "",
    ]
    lines.extend(f"- {channel}：{count} 个文件" for channel, count in sorted(counts.items()))
    lines.extend(
        [
            "",
            "## 质量与追溯",
            "",
            "- `05_资源索引/统一交付资源索引.xlsx` 可筛选全部文件。",
            "- JSON 索引记录每个文件的 SHA256，可用于校验传输完整性。",
            "- 原始协作路径仍随安装包保留，仅用于技术追溯。",
        ]
    )
    (TARGET / "README.md").write_text("\n".join(lines), encoding="utf-8")


def _write_acceptance(records: list[dict[str, object]]) -> None:
    required_channels = {"01_图文翻译", "02_术语与风格", "03_DOCX翻译", "04_音视频翻译"}
    actual_channels = {str(record["任务通道"]) for record in records}
    wav_files = [record for record in records if str(record["格式"]) == "WAV"]
    acceptance = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "passed": required_channels.issubset(actual_channels) and bool(wav_files),
        "file_count": len(records),
        "total_size_bytes": sum(int(record["大小Bytes"]) for record in records),
        "channels": sorted(actual_channels),
        "required_channels": sorted(required_channels),
        "generated_wav_count": len(wav_files),
        "catalog_has_sha256": all(bool(record["SHA256"]) for record in records),
    }
    path = TARGET / "05_资源索引/最终验收.json"
    path.write_text(json.dumps(acceptance, ensure_ascii=False, indent=2), encoding="utf-8")
    if not acceptance["passed"]:
        raise RuntimeError(f"Unified delivery acceptance failed: {path}")


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


if __name__ == "__main__":
    raise SystemExit(main())
